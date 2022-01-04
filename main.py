# -*- encoding: utf-8 -*-

import difflib
import os
import sys

import openpyxl
import xlrd
from configparser import ConfigParser
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 新Excel表格的路径
new_table = None
new_file_name = ""

# 旧Excel表格的路径
old_table = None
old_file_name = ""

# 导出的excel表格的路径
output_file_name = ""

# 重复率阈值
max_rate = 0


def process_config():
    global new_table
    global old_table
    global new_file_name
    global old_file_name
    global output_file_name
    global max_rate

    # 开发用
    # config_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(sys.executable)))) + '\config.ini'
    # 生产用
    config_path = os.path.dirname(os.path.realpath(sys.executable)) + '\config.ini'
    print config_path
    if not os.path.exists(config_path):
        print("No config.ini")
    cp = ConfigParser()
    cp.read(config_path, encoding='utf-8-sig')
    new_file_name = cp.get('config', 'new_file_name')
    # new_file_name = new_file_name.decode('utf-8')
    new_data = xlrd.open_workbook(new_file_name)
    new_table = new_data.sheets()[0]

    old_file_name = cp.get('config', 'old_file_name')
    # old_file_name = old_file_name.decode('utf-8')
    old_data = xlrd.open_workbook(old_file_name)
    old_table = old_data.sheets()[0]

    output_file_name = cp.get('config', 'output_file_name')

    max_rate = float(cp.get('config', 'max_rate'))


def import_excel(new_table, old_table):
    # 新增试题索引
    new_index = 0
    # 保存匹配结果
    result = []
    # 保存最大匹配率
    max_match_rate = 0.0
    # 保存最大匹配行
    max_match_rownum = 0

    for new_row in range(new_table.nrows):
        if new_index < 10:
            new_index = new_index + 1
            continue
        new_array = {'题目': new_table.cell_value(new_row, 3), '解析': new_table.cell_value(new_row, 4),
                     '知识点': new_table.cell_value(new_row, 6),
                     '选项A': new_table.cell_value(new_row, 7), '选项B': new_table.cell_value(new_row, 8),
                     '选项C': new_table.cell_value(new_row, 9),
                     '选项D': new_table.cell_value(new_row, 10), '答案': new_table.cell_value(new_row, 5)}

        # 新文件题型
        new_question_type = new_table.cell_value(new_row, 0)

        # 库中实体索引
        old_index = 0

        # 交叉遍历，获取最高匹配率和最高匹配行
        for old_row in range(old_table.nrows):
            if old_index < 10:
                old_index = old_index + 1
                continue
            old_array = {'题目': old_table.cell_value(old_row, 3), '解析': old_table.cell_value(old_row, 4),
                         '知识点': old_table.cell_value(old_row, 6),
                         '选项A': old_table.cell_value(old_row, 7), '选项B': old_table.cell_value(old_row, 8),
                         '选项C': old_table.cell_value(old_row, 9),
                         '选项D': old_table.cell_value(old_row, 10), '答案': old_table.cell_value(old_row, 5)}

            # 匹配率
            diff_rate = 0.0

            if new_question_type == u"判断":
                diff_rate = compute_match_rate(new_array.get("题目"), old_array.get("题目"))
            else:
                # 取5个匹配要素的平均值（题目和4个选项）
                diff_rate1 = compute_match_rate(new_array.get("题目"), old_array.get("题目"))
                diff_rate2 = compute_match_rate(new_array.get("选项A"), old_array.get("选项A"))
                diff_rate3 = compute_match_rate(new_array.get("选项B"), old_array.get("选项B"))
                diff_rate4 = compute_match_rate(new_array.get("选项C"), old_array.get("选项C"))
                diff_rate5 = compute_match_rate(new_array.get("选项D"), old_array.get("选项D"))
                diff_rate = (diff_rate1 + diff_rate2 + diff_rate3 + diff_rate4 + diff_rate5) / 5

            if diff_rate > max_match_rate:
                max_match_rate = diff_rate
                max_match_rownum = old_index + 1

            old_index = old_index + 1

        print ("new_file_row: " + str(new_index + 1) + ", max_match_rate: " + str(
            max_match_rate) + ", max_match_row: " + str(max_match_rownum))
        temp_map = {'new_row': new_index + 1, 'max_rate': str(max_match_rate), 'old_row': str(max_match_rownum)}
        result.append(temp_map)

        max_match_rate = 0.0
        max_match_rownum = 0
        new_index = new_index + 1

    print "Reverse the result...."
    write_excel(result)
    print "Mission Completed"


# 重写excel
def write_excel(result):
    global max_rate
    global output_file_name

    # 定义单元格格式
    font = Font(name='宋体'.decode('utf-8'), size=11, color='00000000');
    alignment = Alignment(horizontal='center', vertical='center')
    fill = PatternFill(patternType='solid', fgColor='00FFEC8B')
    fill_beyond = PatternFill(patternType='solid', fgColor='00FA8072')
    border = Border(left=Side(style='thin', color='00000000'),
                    right=Side(style='thin', color='00000000'),
                    top=Side(style='thin', color='00000000'),
                    bottom=Side(style='thin', color='00000000'))

    wb = openpyxl.load_workbook(new_file_name)
    ws = wb[wb.sheetnames[0]]
    ws.cell(10, 18).value = '最高匹配率'
    ws.cell(10, 18).font = font
    ws.cell(10, 18).alignment = alignment
    ws.cell(10, 18).fill = fill
    ws.cell(10, 18).border = border

    ws.cell(10, 19).value = '最高匹配行'
    ws.cell(10, 19).font = font
    ws.cell(10, 19).alignment = alignment
    ws.cell(10, 19).fill = fill
    ws.cell(10, 19).border = border

    ws.cell(10, 20).value = '是否重复'
    ws.cell(10, 20).font = font
    ws.cell(10, 20).alignment = alignment
    ws.cell(10, 20).fill = fill
    ws.cell(10, 20).border = border

    # 循环遍历往表格插值
    for each in result:
        rate = format(float(each.get('max_rate')) * 100, '.2f')
        ws.cell(each.get('new_row'), 18).value = str(rate) + '%'
        ws.cell(each.get('new_row'), 18).font = font
        ws.cell(each.get('new_row'), 18).alignment = alignment
        ws.cell(each.get('new_row'), 18).border = border
        if max_rate <= float(rate):
            ws.cell(each.get('new_row'), 18).fill = fill_beyond
            ws.cell(each.get('new_row'), 20).value = '是'
            ws.cell(each.get('new_row'), 20).font = font
            ws.cell(each.get('new_row'), 20).alignment = alignment
            ws.cell(each.get('new_row'), 20).border = border

        ws.cell(each.get('new_row'), 19).value = each.get('old_row')
        ws.cell(each.get('new_row'), 19).font = font
        ws.cell(each.get('new_row'), 19).alignment = alignment
        ws.cell(each.get('new_row'), 19).border = border

    wb.save(output_file_name)


# 算两个变量的匹配率
def compute_match_rate(new, old):
    str1 = ''
    str2 = ''
    if isinstance(new, int) or isinstance(new, float) or isinstance(new, int):
        str1 = str(new)
    else:
        str1 = new
    if isinstance(old, int) or isinstance(old, float) or isinstance(old, int):
        str2 = str(old)
    else:
        str2 = old

    return difflib.SequenceMatcher(None, str1, str2).ratio()


if __name__ == '__main__':
    process_config()
    import_excel(new_table, old_table)
